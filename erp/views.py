from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .forms import CustomLoginForm, FichaNavioForm, FichaQuimicoForm, FichaVehiculoForm, FichaHerramientasForm
from .models import FichaNavio, FichaQuimico, FichaVehiculo, FichaHerramientas
import logging

logger = logging.getLogger(__name__)
#---Excel
import openpyxl
from django.http import HttpResponse, JsonResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#---------------------------------------------------

def home(request):
    return render(request, 'html/home.html')

#--------------Login
def login_view(request):
    if request.user.is_authenticated:
        return redirect('erp:menu')

    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('erp:menu')
            else:
                messages.error(request, "Correo o contraseña incorrectos. Por favor, inténtalo de nuevo.")

    form = CustomLoginForm()
    return render(request, 'html/login.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('erp:login')

#-----------------Menu
@login_required
def menu_view(request):
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"
    return render(request, 'html/menu.html', {'nombre_usuario': nombre_usuario})

#-----------------Inventario
@login_required
def inventario(request):
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"
    return render(request, 'html/MenuInventario.html', {'nombre_usuario': nombre_usuario})

#-----------------Operaciones
@login_required
def gestorOperaciones(request):
    fichas = FichaNavio.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Creación: {ficha.fecha_creacion}")
    
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaNavioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de Navio guardada exitosamente.')
            return redirect('erp:gestor-operaciones')
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaNavioForm()

    return render(request, 'html/gestorOperaciones.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def eliminar_ficha(request, ficha_id):
    ficha = get_object_or_404(FichaNavio, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de Navio eliminada exitosamente.')
    return redirect('erp:gestor-operaciones')

@login_required
def nueva_ficha(request):
    if request.method == 'POST':
        form = FichaNavioForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fecha_creacion = timezone.now().date()
            ficha.save()
            messages.success(request, 'Ficha de Navio guardada exitosamente.')
            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-operaciones')
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')

    return render(request, 'html/fichaNavio.html')

def descargar_excel(request):
    # Obtener datos de la base de datos
    fichas = FichaNavio.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["Nave", "Viaje", "Puerto", "Carga", "Procedencia", "Tipo de Servicio",
               "Armador", "Agencia", "Proximo Puerto", "Encalado", "ETA", "Hora Registro ETA",
               "Bomba Sumergible", "Cubierta", "Shape Box", "PCR", "Hora Registro PCR",
               "Fecha Creacion", "Cantidad de Personas", "Puerto"]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.Nave
        ws[f"B{row_num}"] = ficha.Viaje
        ws[f"C{row_num}"] = ficha.Puerto
        ws[f"D{row_num}"] = ficha.Carga
        ws[f"E{row_num}"] = ficha.Procedencia
        ws[f"F{row_num}"] = ficha.TipoServicio
        ws[f"G{row_num}"] = ficha.Armador
        ws[f"H{row_num}"] = ficha.Agencia
        ws[f"I{row_num}"] = ficha.ProximoPuerto
        ws[f"J{row_num}"] = ficha.Encalado
        ws[f"K{row_num}"] = ficha.ETA.strftime('%d-%m') if ficha.ETA else None
        ws[f"L{row_num}"] = ficha.horaRegistroETA.strftime('%H:%M') if ficha.horaRegistroETA else None
        ws[f"M{row_num}"] = ficha.Bombasumergible
        ws[f"N{row_num}"] = ficha.Cubierta
        ws[f"O{row_num}"] = ficha.ShapeBox
        ws[f"P{row_num}"] = ficha.PCR
        ws[f"Q{row_num}"] = ficha.horaRegistroPCR.strftime('%H:%M') if ficha.horaRegistroPCR else None
        ws[f"R{row_num}"] = ficha.fecha_creacion.strftime('%d-%m-%Y %H:%M:%S')
        ws[f"S{row_num}"] = ficha.cantidadPersonas
        ws[f"T{row_num}"] = ficha.puerto

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_navio.xlsx'
    wb.save(response)

    return response


#-------------Quimicos
def gestorQuimico(request):  
    fichas = FichaQuimico.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Registro: {ficha.fecha_registro}")
    
    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaQuimicoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de Quimicos guardada exitosamente.')
            return redirect('erp:gestor-operaciones')
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaQuimicoForm()

    return render(request, 'html/gestorQuimico.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaQuimico(request):
    if request.method == 'POST':
        form = FichaQuimicoForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fecha_registro = timezone.now().date()
            print("Valores del formulario antes de guardar:", form.cleaned_data)

            ficha.save()

            messages.success(request, 'Ficha de Quimico guardada exitosamente.')
            
            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-quimico')  # Redirige a la página 'gestor-quimico'
        else:
            print("Errores de validación del formulario:", form.errors)
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaQuimicoForm()

    return render(request, 'html/fichaQuimico.html', {'form': form})

def eliminar_fichaQuimico(request, ficha_id):
    ficha = get_object_or_404(FichaQuimico, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de quimico eliminada exitosamente.')
    return redirect('erp:gestor-quimico')

def descargar_excelQuimico(request):
    # Obtener datos de la base de datos
    fichas = FichaQuimico.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["tipo_quimico", "fecha_registro", "capacidad_bines", "lugar_almacenamiento"]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.tipo_quimico
        ws[f"B{row_num}"] = ficha.fecha_registro
        ws[f"C{row_num}"] = ficha.capacidad_bines
        ws[f"D{row_num}"] = ficha.lugar_almacenamiento

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_quimica.xlsx'
    wb.save(response)

    return response


#-------Vehiculo
def gestorVehiculo(request):
    fichas = FichaVehiculo.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaVehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de Vehículo guardada exitosamente.')
            return redirect('erp:gestor-vehiculo')  # Redirige a la página 'gestor-quimico'
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/gestorVehiculo.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaVehiculo(request):
    if request.method == 'POST':
        form = FichaVehiculoForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fecha_ingreso = timezone.now().date()
            print("Valores del formulario antes de guardar:", form.cleaned_data)

            ficha.save()

            messages.success(request, 'Ficha de Vehiculo guardada exitosamente.')
            
            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-vehiculo')  # Redirige a la página 'gestor-quimico'
        else:
            print("Errores de validación del formulario:", form.errors)
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/fichaVehiculo.html', {'form': form})

def eliminar_fichaVehiculo(request, ficha_id):
    ficha = get_object_or_404(FichaVehiculo, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de vehiculo eliminada exitosamente.')
    return redirect('erp:gestor-vehiculo')

def descargar_excelVehiculo(request):
    # Obtener datos de la base de datos
    fichas = FichaVehiculo.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["marca", "modelo", "patente", "fecha_ingreso", "chasis", "tipo_vehiculo", "tipo_combustible"]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.marca
        ws[f"B{row_num}"] = ficha.modelo
        ws[f"C{row_num}"] = ficha.patente
        ws[f"D{row_num}"] = ficha.fecha_ingreso
        ws[f"E{row_num}"] = ficha.chasis
        ws[f"F{row_num}"] = ficha.tipo_vehiculo
        ws[f"G{row_num}"] = ficha.tipo_combustible

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_vehiculos.xlsx'
    wb.save(response)

    return response

#-------Herramientas
@login_required
def gestorHerramientas(request):
    fichas = FichaHerramientas.objects.all()
    for ficha in fichas:
        print(f"Ficha: {ficha}, Fecha de Ingreso: {ficha.fecha_ingreso}")

    nombre_usuario = request.user.nombre if request.user.is_authenticated else "Invitado"

    if request.method == 'POST':
        form = FichaVehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ficha de herramientas guardada exitosamente.')
            return redirect('erp:gestor-herramientas')  # Redirige a la página 'gestor-quimico'
        else:
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/gestorHerramientas.html', {'form': form, 'fichas': fichas, 'nombre_usuario': nombre_usuario})

@login_required
def nuevaFichaHerramientas(request):
    if request.method == 'POST':
        form = FichaHerramientasForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.fechaIngreso = timezone.now().date()
            print("Valores del formulario antes de guardar:", form.cleaned_data)

            ficha.save()

            messages.success(request, 'Ficha de herramientas guardada exitosamente.')
            
            print(f"Ficha guardada: {ficha}")

            # Redirige directamente desde Django
            return redirect('erp:gestor-herramientas')  # Redirige a la página 'gestor-quimico'
        else:
            print("Errores de validación del formulario:", form.errors)
            messages.error(request, 'Error en el formulario. Por favor, verifica los datos.')
    else:
        form = FichaVehiculoForm()

    return render(request, 'html/fichaHerramienta.html', {'form': form})
def descargar_excelHerramientas(request):
    # Obtener datos de la base de datos
    fichas = FichaHerramientas.objects.all()

    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agregar encabezados a la primera fila
    headers = ["marca", "fecha_ingreso", "modelo", "cantidad_herramientas", "tipo_herramienta"]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Agregar datos a las filas siguientes
    for row_num, ficha in enumerate(fichas, 2):
        ws[f"A{row_num}"] = ficha.marca
        ws[f"B{row_num}"] = ficha.fecha_ingreso
        ws[f"C{row_num}"] = ficha.modelo
        ws[f"D{row_num}"] = ficha.cantidad_herramientas
        ws[f"E{row_num}"] = ficha.tipo_herramienta

    # Crear una respuesta de archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fichas_herramientas.xlsx'
    wb.save(response)

    return response

def eliminar_fichaHerramientas(request, ficha_id):
    ficha = get_object_or_404(FichaHerramientas, id=ficha_id)
    ficha.delete()
    messages.success(request, 'Ficha de herramientas eliminada exitosamente.')
    return redirect('erp:gestor-herramientas')
